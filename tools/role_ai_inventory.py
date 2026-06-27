#!/usr/bin/env python3
"""Rol/accountability mens-vs-AI inventarisatie over meerdere organisaties.

Leest GlassFrog governance-export-PDF's (NL of EN) uit een map, één (of meer) PDF per organisatie,
inventariseert per accountability of die door een AI-agent kan (AI), samen met de mens (Hybride),
of door de mens moet (Mens), en schrijft een Excel met:
  - blad "Vergelijking": per organisatie de procentuele verdeling + het gemiddelde + de afwijking
    van elke organisatie t.o.v. dat gemiddelde;
  - per organisatie een detailblad met alle accountabilities en hun classificatie.

Facilitator- en Secretaris/Secretary-rollen worden uitgesloten (procesrollen, tellen niet mee).

De classificatie is een keyword-heuristiek (consistente eerste inschatting, te verfijnen).
Hetzelfde recept draait over alle organisaties, dus de vergelijking is apples-to-apples.

Gebruik:
    python tools/role_ai_inventory.py <map_met_pdfs> <output.xlsx>
"""
from __future__ import annotations
import os
import re
import sys

from pypdf import PdfReader
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

# Lange exportnamen → korte, nette organisatielabels.
ORG_LABEL = {
    "onlineveilingmeester.nl": "OVM",
    "onlineveilingmeester": "OVM",
    "nooch footwear": "Nooch",
    "findyourgroup": "Findyourgroup",
    "obelink": "Obelink",
}
EXCLUDE_ROLES = {"facilitator", "secretary", "secretaris"}

# Sectiekoppen (NL + EN). 'Purpose' is in beide exports de record-grens.
_ACC_HEADERS = ("Verantwoordelijkheden", "Accountabilities")
_STOP_HEADERS = ("Projecten", "Projects", "Domeinen", "Domains", "Role Fillers",
                 "Rolvervullers", "Leden", "Members", "Purpose")
_NONE_ACC = ("Er zijn geen verantwoordelijkheden", "There are no Accountabilities")

# ── classificatie ──
_MENS = [r'relatie', r'relationship', r'vertrouwen', r'\btrust', r'onderhandel', r'negoti',
         r'juridisch', r'\blegal', r'aansprakelijk', r'liabilit', r'onderteken', r'handtekening',
         r'\bsign(ing|ature|ed)?\b', r'investeer', r'investor', r'fundrais', r'vertegenwoordig',
         r'\brepresent', r'leidinggev', r'aansturen', r'\bcoach', r'\bwerven\b', r'recruit',
         r'\bhiring\b', r'ontslag', r'ontslaan', r'\bbenoemen', r'aanstellen', r'cultuur',
         r'\bculture', r'\bbezoek', r'\bvisit', r'fysiek', r'physical', r'\bconsent',
         r'eindverantwoord', r'functionerings', r'beoordelingsgesprek', r'sollicit', r'\bnetwerk',
         r'\bnetwork', r'stakeholder', r'partnerschap', r'partnership', r'klantrelatie',
         r'motiveren', r'inspireren', r'\bmentor', r'welzijn', r'well-?being', r'personeel']
_MENS_RE = [re.compile(p) for p in _MENS]
_AI_VERBS = {
    "monitoren", "bewaken", "bijhouden", "rapporteren", "analyseren", "verzamelen", "plannen",
    "inplannen", "agenderen", "vastleggen", "notuleren", "documenteren", "publiceren",
    "signaleren", "tracken", "archiveren", "administreren", "interpreteren", "opstellen",
    "controleren", "delen", "verwerken", "registreren", "genereren", "samenvatten", "updaten",
    "bijwerken", "opvolgen", "factureren",
    "monitoring", "tracking", "reporting", "analyzing", "collecting", "gathering", "scheduling",
    "capturing", "recording", "documenting", "publishing", "flagging", "drafting", "preparing",
    "checking", "verifying", "sharing", "processing", "generating", "summarizing", "updating",
}


def classify(acc: str):
    low = acc.lower()
    for rx in _MENS_RE:
        if rx.search(low):
            return ("Mens", "-", "Relationeel/juridisch/zeggenschap: blijft bij de rolhouder.")
    words = re.findall(r"[a-zäöüéè]+", low)
    first = words[0] if words else ""
    if first in _AI_VERBS:
        return ("AI", "Nu", f"Data-/proceswerk ('{first}'): te automatiseren door een agent.")
    return ("Hybride", "Begeleid", "AI doet het voorwerk; de rolhouder beslist of doet het "
            "relationele/fysieke deel.")


# ── parsing ──
def _org_from_filename(fn: str) -> str:
    stem = os.path.splitext(os.path.basename(fn))[0]
    base = re.split(r"governance", stem, flags=re.IGNORECASE)[0]
    base = re.sub(r"[-–]\s*\d.*$", "", base)          # datum/volgnummer eraf
    base = base.replace("-proces", "").strip(" -–_")
    key = base.lower().strip()
    return ORG_LABEL.get(key, base or stem)


def _clean_lines(text: str) -> list[str]:
    out = []
    for l in text.splitlines():
        s = l.replace("\x0c", "").strip()
        if not s:
            continue
        if re.match(r'^\d{4}-\d{2}-\d{2}.*UTC$', s) or re.match(r'^\d{1,3}$', s):
            continue
        if re.match(r'^\(https?://', s):          # glassfrog-link-regels (continuatie-ruis)
            continue
        out.append(s)
    return out


def parse_export(text: str) -> list[dict]:
    """Geef records terug: {name, parent, accs[]}. Facilitator/Secretaris uitgesloten."""
    lines = _clean_lines(text)
    purpose_idx = [i for i, l in enumerate(lines) if l in ("Purpose",)]
    recs = []
    for k, pi in enumerate(purpose_idx):
        # Naam/ouder robuust voor beide export-formaten:
        #  - NL (pdftotext): <naam>, (Ouder), Purpose   → naam = pi-2, ouder = pi-1
        #  - EN (pypdf):     (Ouder), <naam>, Purpose   → naam = pi-1, ouder ervoor
        prev = lines[pi-1] if pi > 0 else ""
        m = re.match(r'^\((.+)\)$', prev)
        if m:
            parent = m.group(1)
            name = lines[pi-2] if pi >= 2 else "?"
        else:
            name = prev or "?"
            parent = "—"
            for j in range(pi-2, max(pi-5, -1), -1):
                mm = re.match(r'^\((.+)\)$', lines[j])
                if mm:
                    parent = mm.group(1); break
        if re.match(r'^\(.+\)$', name):
            name = "?"
        if name.strip().lower() in EXCLUDE_ROLES:
            continue
        end = purpose_idx[k+1]-1 if k+1 < len(purpose_idx) else len(lines)
        block = lines[pi:end]
        accs = []
        hdr = next((h for h in _ACC_HEADERS if h in block), None)
        if hdr:
            seg = block[block.index(hdr)+1:]
            cut = len(seg)
            for st in _STOP_HEADERS:
                if st in seg:
                    cut = min(cut, seg.index(st))
            seg = seg[:cut]
            if seg and not any(seg[0].startswith(n) for n in _NONE_ACC):
                cur = ""
                for ln in seg:
                    new = bool(re.match(r'^[A-ZÄÖÜ]', ln)) and cur
                    if new and len(ln.split()) <= 2 and not cur.rstrip().endswith('.'):
                        new = False
                    if new:
                        accs.append(cur.strip()); cur = ln
                    else:
                        cur = (cur + " " + ln).strip() if cur else ln
                if cur.strip():
                    accs.append(cur.strip())
        recs.append({"name": name, "parent": parent, "accs": accs})
    return recs


# ── Excel ──
FONT = "Arial"
FILLS = {"AI": "C6EFCE", "Hybride": "FFEB9C", "Mens": "BDD7EE"}
FCOL = {"AI": "006100", "Hybride": "9C6500", "Mens": "1F4E78"}
_thin = Side(style="thin", color="DDD4C0")
BORDER = Border(left=_thin, right=_thin, top=_thin, bottom=_thin)


def _safe_sheet(name: str, used: set) -> str:
    s = re.sub(r'[\[\]\*\?:/\\]', '', name)[:28] or "org"
    base, n = s, 1
    while s in used:
        s = f"{base}_{n}"; n += 1
    used.add(s)
    return s


def build(orgs: list[tuple[str, list[list]]], out: str):
    """orgs: lijst (label, rows) waarbij rows = [cirkel, rol, accountability, cat, rijpheid, waarom]."""
    wb = Workbook()
    comp = wb.active; comp.title = "Vergelijking"
    used = {"Vergelijking", "Methode"}
    sheet_of = {}

    # detailbladen
    for label, rows in orgs:
        sn = _safe_sheet(label, used); sheet_of[label] = sn
        ws = wb.create_sheet(sn)
        ws.append(["Cirkel", "Rol", "Accountability", "Uitvoerder", "AI-rijpheid", "Waarom"])
        for c in range(1, 7):
            cell = ws.cell(1, c); cell.fill = PatternFill("solid", fgColor="16323A")
            cell.font = Font(name=FONT, bold=True, color="FFFFFF")
        for r in rows:
            ws.append(r); i = ws.max_row
            ec = ws.cell(i, 4); ec.fill = PatternFill("solid", fgColor=FILLS[r[3]])
            ec.font = Font(name=FONT, bold=True, color=FCOL[r[3]]); ec.alignment = Alignment(horizontal="center")
            for c in range(1, 7):
                cl = ws.cell(i, c); cl.border = BORDER
                if c != 4:
                    cl.font = Font(name=FONT)
                cl.alignment = Alignment(vertical="top", wrap_text=(c in (3, 6)),
                                         horizontal="center" if c == 5 else "left")
        for col, w in {"A": 20, "B": 26, "C": 60, "D": 11, "E": 12, "F": 50}.items():
            ws.column_dimensions[col].width = w
        ws.freeze_panes = "A2"; ws.auto_filter.ref = f"A1:F{ws.max_row}"

    # vergelijkingsblad
    comp["A1"] = "Mens vs. AI per accountability — vergelijking tussen organisaties"
    comp["A1"].font = Font(name=FONT, bold=True, size=14)
    comp["A2"] = ("Facilitator- en Secretaris-rollen uitgesloten. Classificatie = keyword-heuristiek "
                  "(consistent over alle organisaties); eerste inschatting, te verfijnen.")
    comp["A2"].font = Font(name=FONT, italic=True, color="9C6500")
    comp.merge_cells("A2:F2"); comp["A2"].alignment = Alignment(wrap_text=True); comp.row_dimensions[2].height = 28

    hdr = ["Organisatie", "# accountabilities", "AI %", "Hybride %", "Mens %"]
    comp.append([])
    comp.append(hdr)
    hrow = comp.max_row
    for c in range(1, 6):
        cell = comp.cell(hrow, c); cell.fill = PatternFill("solid", fgColor="16323A")
        cell.font = Font(name=FONT, bold=True, color="FFFFFF")

    first_data = hrow + 1
    for label, rows in orgs:
        sn = sheet_of[label]
        col = f"'{sn}'!$D$2:$D$100000"
        comp.append([label,
                     f'=COUNTA({col})',
                     f'=COUNTIF({col},"AI")/COUNTA({col})',
                     f'=COUNTIF({col},"Hybride")/COUNTA({col})',
                     f'=COUNTIF({col},"Mens")/COUNTA({col})'])
        i = comp.max_row
        for c in (3, 4, 5):
            comp.cell(i, c).number_format = "0%"
        for c in range(1, 6):
            comp.cell(i, c).font = Font(name=FONT); comp.cell(i, c).border = BORDER
    last_data = comp.max_row
    # gemiddelde-rij (gemiddelde van de organisatie-percentages, gelijk gewogen)
    comp.append(["Gemiddelde",
                 f'=AVERAGE(B{first_data}:B{last_data})',
                 f'=AVERAGE(C{first_data}:C{last_data})',
                 f'=AVERAGE(D{first_data}:D{last_data})',
                 f'=AVERAGE(E{first_data}:E{last_data})'])
    avg_row = comp.max_row
    comp.cell(avg_row, 2).number_format = "0.0"
    for c in (3, 4, 5):
        comp.cell(avg_row, c).number_format = "0%"
    for c in range(1, 6):
        comp.cell(avg_row, c).font = Font(name=FONT, bold=True)
        comp.cell(avg_row, c).fill = PatternFill("solid", fgColor="F1ECDF")
        comp.cell(avg_row, c).border = BORDER

    # afwijking t.o.v. gemiddelde (procentpunten)
    comp.append([]); comp.append(["Afwijking t.o.v. gemiddelde (procentpunten)"])
    comp.cell(comp.max_row, 1).font = Font(name=FONT, bold=True, size=11)
    comp.append(["Organisatie", "Δ AI", "Δ Hybride", "Δ Mens"])
    dh = comp.max_row
    for c in range(1, 5):
        comp.cell(dh, c).fill = PatternFill("solid", fgColor="16323A")
        comp.cell(dh, c).font = Font(name=FONT, bold=True, color="FFFFFF")
    for n, (label, _rows) in enumerate(orgs):
        src = first_data + n
        comp.append([label,
                     f'=C{src}-$C${avg_row}', f'=D{src}-$D${avg_row}', f'=E{src}-$E${avg_row}'])
        i = comp.max_row
        for c in (2, 3, 4):
            cell = comp.cell(i, c); cell.number_format = "+0%;-0%;0%"
            cell.font = Font(name=FONT); cell.border = BORDER
        comp.cell(i, 1).font = Font(name=FONT); comp.cell(i, 1).border = BORDER

    for col, w in {"A": 22, "B": 16, "C": 12, "D": 12, "E": 10}.items():
        comp.column_dimensions[col].width = w

    # methode-blad
    mt = wb.create_sheet("Methode")
    notes = [
        ("Methode", True, 13),
        ("Bron: GlassFrog governance-export per organisatie (NL of EN).", False, 11),
        ("Facilitator- en Secretaris/Secretary-rollen zijn uitgesloten (procesrollen).", False, 11),
        ("", False, 11),
        ("Classificatie per accountability:", True, 11),
        ("AI = data-/proceswerk dat een agent binnen de rol autonoom kan doen.", False, 11),
        ("Hybride = de agent doet het voorwerk; de mens (rolhouder) beslist of voert het", False, 11),
        ("   relationele/fysieke deel uit.", False, 11),
        ("Mens = relationeel, juridisch, fysiek of zeggenschap; blijft bij de rolhouder.", False, 11),
        ("", False, 11),
        ("De classificatie is een keyword-heuristiek: een consistente eerste inschatting,", False, 11),
        ("bedoeld om samen te verfijnen. 'Hybride' is de vergaarbak; 'Mens' is een ondergrens.", False, 11),
        ("Omdat hetzelfde recept over alle organisaties draait, is de vergelijking eerlijk.", False, 11),
        ("", False, 11),
        ("Visie: de rol blijft van een mens; AI-agents pakken binnen de rol specifieke", False, 11),
        ("accountabilities op. Dit overzicht is de kaart daarvoor.", False, 11),
    ]
    for r, (txt, bold, size) in enumerate(notes, start=1):
        mt.cell(r, 1, txt).font = Font(name=FONT, bold=bold, size=size)
    mt.column_dimensions["A"].width = 95

    wb.save(out)


def main():
    if len(sys.argv) < 3:
        print("gebruik: python tools/role_ai_inventory.py <map_met_pdfs> <output.xlsx>")
        sys.exit(1)
    indir, out = sys.argv[1], sys.argv[2]
    files = sorted(f for f in os.listdir(indir) if f.lower().endswith((".pdf", ".txt")))
    if not files:
        print("geen PDF's/TXT's gevonden in", indir); sys.exit(1)
    grouped: dict[str, list[dict]] = {}
    for fn in files:
        label = _org_from_filename(fn)
        path = os.path.join(indir, fn)
        if fn.lower().endswith(".txt"):
            text = open(path, encoding="utf-8", errors="ignore").read()
        else:
            text = "\n".join(p.extract_text() or "" for p in PdfReader(path).pages)
        grouped.setdefault(label, []).extend(parse_export(text))
    orgs = []
    for label, recs in grouped.items():
        rows = []
        for r in recs:
            for a in r["accs"]:
                cat, ripe, why = classify(a)
                rows.append([r["parent"], r["name"], a, cat, ripe, why])
        if rows:
            orgs.append((label, rows))
            n = len(rows)
            ai = sum(1 for x in rows if x[3] == "AI")
            hy = sum(1 for x in rows if x[3] == "Hybride")
            me = sum(1 for x in rows if x[3] == "Mens")
            print(f"{label:16s} acc={n:4d}  AI={ai/n:4.0%}  Hybride={hy/n:4.0%}  Mens={me/n:4.0%}")
    build(orgs, out)
    print("\ngeschreven:", out, "| organisaties:", len(orgs))


if __name__ == "__main__":
    main()
